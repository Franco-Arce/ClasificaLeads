import streamlit as st
import json
import pandas as pd
from logic import process_data
import docx

st.set_page_config(page_title="Lead Classifier", layout="wide")

st.title("📊 Lead Classifier & Analyzer")
st.markdown("""
Sube tu archivo JSON (o .docx con JSON) de logs de chat para procesarlo y clasificar los leads en **SPAM**, **MQL** o **SQL** 
basado en el sistema de scoring definido.

### Sistema de Scoring
| Categoría | Puntos Máximos |
|-----------|----------------|
| Motivación del Lead | 40 pts |
| Intención de Pago | 30 pts |
| Comportamiento/Timing | 30 pts |

**Clasificación:** SPAM (0 pts) | MQL (1-49 pts) | SQL (50-100 pts)
""")

uploaded_file = st.file_uploader("Cargar archivo de Chat Logs (JSON/DOCX)", type=["json", "docx"])
neotel_file = st.file_uploader("Cargar base Neotel (Excel) - Opcional", type=["xls", "xlsx"])

if uploaded_file is not None:
    try:
        data = None
        # Determine file type and load content
        if uploaded_file.name.endswith('.json'):
            data = json.load(uploaded_file)
        elif uploaded_file.name.endswith('.docx'):
            doc = docx.Document(uploaded_file)
            full_text = []
            for para in doc.paragraphs:
                full_text.append(para.text)
            json_text = "\n".join(full_text)
            data = json.loads(json_text)
            
        if data is None:
             st.error("No se pudo leer el archivo.")
        elif "items" not in data:
            st.error("El JSON no tiene el formato correcto (falta la clave 'items').")
        else:
            st.success(f"Archivo de logs cargado correctamente. {len(data['items'])} mensajes encontrados.")
            
            neotel_df = None
            if neotel_file is not None:
                try:
                    neotel_df = pd.read_excel(neotel_file)
                    st.success(f"Base Neotel cargada correctamente. {len(neotel_df)} registros.")
                except Exception as e:
                    st.error(f"Error al leer el archivo Excel de Neotel: {e}")

            if st.button("Procesar Leads"):
                with st.spinner("Procesando conversaciones..."):
                    # Process data
                    results = process_data(data, neotel_df)
                    
                    # Convert to DataFrame for display
                    df = pd.DataFrame(results)
                    
                    # Metrics
                    total_leads = len(df)
                    spam_leads = len(df[df['clasificacion'] == 'SPAM'])
                    sql_leads = len(df[df['clasificacion'] == 'SQL'])
                    mql_leads = len(df[df['clasificacion'] == 'MQL'])
                    
                    # Calculate average score for non-SPAM leads
                    non_spam_df = df[df['clasificacion'] != 'SPAM']
                    avg_score = non_spam_df['score_total'].mean() if len(non_spam_df) > 0 else 0
                    
                    # Display metrics in columns
                    st.subheader("📈 Resumen")
                    col1, col2, col3, col4, col5 = st.columns(5)
                    col1.metric("Total Leads", total_leads)
                    col2.metric("🚫 SPAM", spam_leads, delta=None)
                    col3.metric("📧 MQL", mql_leads)
                    col4.metric("🎯 SQL", sql_leads)
                    col5.metric("📊 Score Promedio", f"{avg_score:.1f}")
                    
                    # Score distribution chart
                    if len(non_spam_df) > 0:
                        st.subheader("📊 Distribución de Scores")
                        
                        col_chart1, col_chart2 = st.columns(2)
                        
                        with col_chart1:
                            # Classification pie chart
                            class_counts = df['clasificacion'].value_counts()
                            st.bar_chart(class_counts)
                        
                        with col_chart2:
                            # Score breakdown averages
                            st.write("**Promedios por Categoría de Score:**")
                            avg_motivation = non_spam_df['score_motivacion'].mean()
                            avg_payment = non_spam_df['score_pago'].mean()
                            avg_behavior = non_spam_df['score_comportamiento'].mean()
                            
                            score_data = pd.DataFrame({
                                'Categoría': ['Motivación (max 40)', 'Pago (max 30)', 'Comportamiento (max 30)'],
                                'Promedio': [avg_motivation, avg_payment, avg_behavior]
                            })
                            st.dataframe(score_data, hide_index=True)
                    
                    # Display Data
                    st.subheader("📋 Resultados Detallados")
                    
                    # Reorder columns for better display
                    display_columns = [
                        'chat_id', 'telefono', 'clasificacion', 'score_total',
                        'score_motivacion', 'score_pago', 'score_comportamiento',
                        'razon_principal', 'señales_clave', 'estado_conversacion',
                        'duracion_chat', 'mensajes_usuario',
                        'utm_source', 'utm_medium', 'utm_origen', 'programa_interes'
                    ]
                    
                    # Only show columns that exist
                    available_columns = [col for col in display_columns if col in df.columns]
                    df_display = df[available_columns]
                    
                    # Interactive Table with Column Config
                    st.dataframe(
                        df_display, 
                        use_container_width=True,
                        column_config={
                            "chat_id": "Chat ID",
                            "telefono": "Teléfono",
                            "clasificacion": st.column_config.TextColumn(
                                "Clasificación",
                                help="SPAM: Descartado, SQL: Sales Qualified Lead, MQL: Marketing Qualified Lead",
                                width="medium"
                            ),
                            "score_total": st.column_config.ProgressColumn(
                                "Score Total",
                                help="Puntaje total (0-100)",
                                format="%d pts",
                                min_value=0,
                                max_value=100,
                            ),
                            "score_motivacion": st.column_config.NumberColumn(
                                "Motivación",
                                help="Puntaje de motivación (max 40)",
                                format="%d pts"
                            ),
                            "score_pago": st.column_config.NumberColumn(
                                "Pago",
                                help="Puntaje de intención de pago (max 30)",
                                format="%d pts"
                            ),
                            "score_comportamiento": st.column_config.NumberColumn(
                                "Comportamiento",
                                help="Puntaje de comportamiento/timing (max 30)",
                                format="%d pts"
                            ),
                            "razon_principal": "Razón",
                            "señales_clave": "Señales Detectadas",
                            "estado_conversacion": "Estado",
                            "duracion_chat": "Duración",
                            "mensajes_usuario": "Msgs Usuario",
                            "utm_source": "UTM Source",
                            "utm_medium": "UTM Medium",
                            "utm_origen": "UTM Origen",
                            "programa_interes": "Programa Interés"
                        }
                    )
                    
                    # Filters
                    st.subheader("🔍 Filtrar Resultados")
                    filter_col1, filter_col2 = st.columns(2)
                    
                    with filter_col1:
                        selected_class = st.multiselect(
                            "Filtrar por Clasificación:",
                            options=['SQL', 'MQL', 'SPAM'],
                            default=['SQL', 'MQL', 'SPAM']
                        )
                    
                    with filter_col2:
                        min_score = st.slider("Score mínimo:", 0, 100, 0)
                    
                    filtered_df = df_display[
                        (df['clasificacion'].isin(selected_class)) & 
                        (df['score_total'] >= min_score)
                    ]
                    
                    if len(filtered_df) != len(df_display):
                        st.write(f"**Mostrando {len(filtered_df)} de {len(df_display)} leads**")
                        st.dataframe(filtered_df, use_container_width=True)
                    
                    # Download Buttons
                    st.subheader("📥 Descargar Resultados")
                    col_d1, col_d2 = st.columns(2)
                    
                    # JSON Download
                    json_output = json.dumps(results, indent=2, ensure_ascii=False)
                    col_d1.download_button(
                        label="📥 Descargar JSON",
                        data=json_output,
                        file_name="leads_clasificados.json",
                        mime="application/json"
                    )
                    
                    # Excel Download with styling
                    import io
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    from openpyxl.utils import get_column_letter
                    
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Leads')
                        
                        # Get the workbook and worksheet
                        workbook = writer.book
                        ws_leads = writer.sheets['Leads']
                        
                        # Define styles
                        header_font = Font(bold=True, color="FFFFFF", size=11)
                        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        # Classification colors
                        sql_fill = PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")  # Green
                        mql_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")  # Yellow
                        spam_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
                        
                        thin_border = Border(
                            left=Side(style='thin', color='CCCCCC'),
                            right=Side(style='thin', color='CCCCCC'),
                            top=Side(style='thin', color='CCCCCC'),
                            bottom=Side(style='thin', color='CCCCCC')
                        )
                        
                        # Style headers
                        for col_num, cell in enumerate(ws_leads[1], 1):
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = thin_border
                        
                        # Find classification column
                        clasif_col = None
                        for col_num, cell in enumerate(ws_leads[1], 1):
                            if cell.value == 'clasificacion':
                                clasif_col = col_num
                                break
                        
                        # Style data rows
                        for row_num in range(2, ws_leads.max_row + 1):
                            for col_num in range(1, ws_leads.max_column + 1):
                                cell = ws_leads.cell(row=row_num, column=col_num)
                                cell.border = thin_border
                                cell.alignment = Alignment(vertical="center")
                            
                            # Color row based on classification
                            if clasif_col:
                                clasif_value = ws_leads.cell(row=row_num, column=clasif_col).value
                                if clasif_value == 'SQL':
                                    for col_num in range(1, ws_leads.max_column + 1):
                                        ws_leads.cell(row=row_num, column=col_num).fill = sql_fill
                                elif clasif_value == 'MQL':
                                    for col_num in range(1, ws_leads.max_column + 1):
                                        ws_leads.cell(row=row_num, column=col_num).fill = mql_fill
                                elif clasif_value == 'SPAM':
                                    for col_num in range(1, ws_leads.max_column + 1):
                                        ws_leads.cell(row=row_num, column=col_num).fill = spam_fill
                        
                        # Auto-fit column widths
                        for col_num in range(1, ws_leads.max_column + 1):
                            max_length = 0
                            column_letter = get_column_letter(col_num)
                            for row in ws_leads.iter_rows(min_col=col_num, max_col=col_num):
                                for cell in row:
                                    try:
                                        if cell.value:
                                            max_length = max(max_length, len(str(cell.value)))
                                    except:
                                        pass
                            adjusted_width = min(max_length + 2, 50)
                            ws_leads.column_dimensions[column_letter].width = adjusted_width
                        
                        # Freeze header row
                        ws_leads.freeze_panes = 'A2'
                        
                        # Create Summary sheet with styling
                        summary_data = {
                            'Métrica': ['Total Leads', 'SPAM', 'MQL', 'SQL', 'Score Promedio'],
                            'Valor': [total_leads, spam_leads, mql_leads, sql_leads, f"{avg_score:.1f}"],
                            'Porcentaje': [
                                '100%',
                                f"{(spam_leads/total_leads*100):.1f}%" if total_leads > 0 else "0%",
                                f"{(mql_leads/total_leads*100):.1f}%" if total_leads > 0 else "0%",
                                f"{(sql_leads/total_leads*100):.1f}%" if total_leads > 0 else "0%",
                                '-'
                            ]
                        }
                        summary_df = pd.DataFrame(summary_data)
                        summary_df.to_excel(writer, index=False, sheet_name='Resumen')
                        
                        ws_summary = writer.sheets['Resumen']
                        
                        # Style summary headers
                        for cell in ws_summary[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = thin_border
                        
                        # Style summary data
                        summary_colors = {
                            'Total Leads': PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
                            'SPAM': spam_fill,
                            'MQL': mql_fill,
                            'SQL': sql_fill,
                            'Score Promedio': PatternFill(start_color="B8D4E3", end_color="B8D4E3", fill_type="solid")
                        }
                        
                        for row_num in range(2, ws_summary.max_row + 1):
                            metric_name = ws_summary.cell(row=row_num, column=1).value
                            fill_color = summary_colors.get(metric_name)
                            for col_num in range(1, ws_summary.max_column + 1):
                                cell = ws_summary.cell(row=row_num, column=col_num)
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                if fill_color:
                                    cell.fill = fill_color
                        
                        # Auto-fit summary columns
                        ws_summary.column_dimensions['A'].width = 20
                        ws_summary.column_dimensions['B'].width = 15
                        ws_summary.column_dimensions['C'].width = 15
                        
                    excel_data = output.getvalue()
                    
                    col_d2.download_button(
                        label="📊 Descargar Excel",
                        data=excel_data,
                        file_name="leads_clasificados.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
    except json.JSONDecodeError:
        st.error("Error al leer el archivo JSON. Asegúrate de que sea un JSON válido.")
    except Exception as e:
        st.error(f"Ocurrió un error inesperado: {e}")
